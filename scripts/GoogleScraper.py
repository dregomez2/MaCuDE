from selenium import webdriver
from bs4 import BeautifulSoup
import time
from bs4.element import Tag
from openpyxl import load_workbook

driver = webdriver.Chrome('./chromedriver')
# Path where results store


URLFile = "FinanceURL.xlsx"
wb = load_workbook(URLFile)
ws = wb['URLs']
# Number of results to record, edit to change # of results
searchNumber=5
university = "MIT"
degree = "Master"
topic = "Finance"
keyword = "Curriculum" 
google_url = "https://www.google.com/search?q=" + university + "+" + degree + "+" + topic + "+" + keyword + "&num=" + str(searchNumber)
driver.get(google_url)
time.sleep(3)

#max_row is a sheet function that gets the last row in a sheet.
newRowLocation = ws.max_row + 1

soup = BeautifulSoup(driver.page_source,'lxml')
result_div = soup.find_all('div', attrs={'class': 'g'})


links = []
titles = []
descriptions = []
for r in result_div:
    # Checks if each element is present, else, raise exception
    try:
        link = r.find('a', href=True)
        title = None
        title = r.find('h3')

        if isinstance(title,Tag):
            title = title.get_text()

        description = None
        description = r.find('span', attrs={'class': 'st'})

        if isinstance(description, Tag):
            description = description.get_text()

        # Check to make sure everything is present before appending
        if link != '' and title != '' and description != '':
            links.append(link['href'])
            ws.cell(column=1,row=newRowLocation, value=topic)
            ws.cell(column=2,row=newRowLocation, value=link['href'])
            ws.cell(column=3,row=newRowLocation, value=university)
            ws.cell(column=4,row=newRowLocation, value=degree)
            ws.cell(column=5,row=newRowLocation, value=keyword)
            ws.cell(column=6,row=newRowLocation, value=title)
            ws.cell(column=7,row=newRowLocation, value=description)
            newRowLocation = newRowLocation + 1
            titles.append(title)
            wb.save(filename=URLFile)   
    # Next loop if one element is not present
    except Exception as e:
        print(e)
        continue

print(links)
wb.close()